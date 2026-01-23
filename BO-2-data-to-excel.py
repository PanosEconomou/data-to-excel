# Vassilis Economou  16/01/2025 v.02
#                   20/01/2026 v.2.1
#                   22/01/2026 v.2.2 (Added Language Toggle)
#                   23/01/2026 v.2.3 (randar)


import openpyxl
from openpyxl import Workbook
import csv
import serial
import serial.tools.list_ports as list_ports
from datetime import datetime
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import warnings
import os
import requests  
import numpy as np  # NEW: Για μετατροπές γωνιών
from itertools import zip_longest


warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")

class SerialDataLogger:
    def __init__(self, root):
        self.root = root
        self.current_lang = "EL" 
        
        self.translations = {
            "EL": {
                "title": "Serial Data Logger [Βασίλης Οικονόμου] v.2.4",
                "setings": "ΡΥΘΜΙΣΕΙΣ",
                "actions": "ΛΕΙΤΟΥΡΓΙΕΣ",
                "instructions": "  Οδηγίες  ",
                "port_label": "Θύρα:",
                "refresh": "Ανανέωση",
                "baud_label": "Baudrate:",
                "file_label": "Αρχείο (.xlsx/.csv):",
                "browse": "Επιλογή",
                "col_titles": "Τίτλοι στηλών:",
                "thingspeak": "ThingSpeak API Key:",
                "ts_interval": "Συχνότητα αποστολής (sec):",
                "sampling": "Καθυστέρηση (ms):",
                "start": "Έναρξη",
                "stop": "Τερματισμός",
                "save": "Αποθήκευση",
                "clear": "Καθαρισμός",
                "graph_win": "Όριο μετρήσεων (Y):",
                "scroll": "Κύληση διαγράμματος",
                "points": "μετρήσεις",
                "log_win": "Kαταγραφή τιμών",
                "listbox_limit": "Όριο γραμμών:",
                "copy": "Αντιγραφή",
                "export_csv": "Εξαγωγή σε .csv",
                "export_xlsx": "Εξαγωγή σε .xlsx",
                "lang_btn": "🇬🇧 English",
                "graph_type": "Τύπος διαγράμματος:", 
                "heading": "Κατεύθυνση",
                "linear_title": "Γραμμική απεικόνιση Δεδομένων",
                "x_label": "Αριθμός μετρήσεων",
                "y_label": "Τιμή",
                "last_points": " τελευταίες"
            },
            "EN": {
                "title": "Serial Data Logger [Vassilis Economou] v.2.4",
                "setings": "SETTINGS",
                "actions": "ACTIONS",
                "instructions": " Instructions ",
                "port_label": "Port:",
                "refresh": "Refresh",
                "baud_label": "Baudrate:",
                "file_label": "File (.xlsx/.csv):",
                "browse": "Browse",
                "col_titles": "Column titles:",
                "thingspeak": "ThingSpeak API Key:",
                "ts_interval": "Interval (sec):",
                "sampling": "Delay (ms):",
                "start": "Start",
                "stop": "Stop",
                "save": "Save",
                "clear": "Clear",
                "graph_win": "Y Limit:",
                "scroll": "Scroll",
                "points": "measurements",
                "log_win": "Data Log",
                "listbox_limit": "Line limit:",
                "copy": "Copy",
                "export_csv": "Export to .csv",
                "export_xlsx": "Export to .xlsx",
                "lang_btn": "🇬🇷 Ελληνικά",
                "graph_type": "Graph type:",
                "heading": "Heading",
                "linear_title": "Linear Data View",
                "x_label": "Number of Measurements",
                "y_label": "Value",
                "last_points": " recent"
            }
        }

        self.root.title(self.translations[self.current_lang]["title"])
        self.root.geometry("1100x750")

        self.serial_port = None
        self.baudrate = tk.IntVar(value=9600)
        self.max_val_limit = tk.IntVar(value=1024)
        self.output_path = tk.StringVar(value=os.path.join(os.getcwd(), "BO_SDL.xlsx"))
        self.times = []
        self.values = []
        self.data_queue = queue.Queue()
        self.stop_event = threading.Event()
        self.sampling_rate = tk.IntVar(value=0)
        self.send_to_thingspeak = tk.BooleanVar(value=False)
        self.thingspeak_api_key = tk.StringVar(value="0J62FHGN0IS42VNQ")
        self.scroll_mode = tk.BooleanVar(value=True)
        self.scroll_window_size = tk.IntVar(value=200)
        self.actual_timestamps = []
        self.listbox_limit = tk.IntVar(value=80000)
        self.ts_interval = tk.IntVar(value=15)
        self.last_ts_send = datetime.min  

        # NEW: Επιλογή τύπου διαγράμματος
        self.graph_type = tk.StringVar(value="Linear")

        self.create_widgets()

    def create_widgets(self):
        t = self.translations[self.current_lang]
        
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        # --- ΑΡΙΣΤΕΡΗ ΠΛΕΥΡΑ ---
        left_side = ttk.Frame(main_frame)
        left_side.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.title_label = ttk.Label(left_side, text=t["title"], font=("Arial", 12, "bold"))
        self.title_label.pack(pady=2)

        self.paned_window = ttk.PanedWindow(left_side, orient=tk.VERTICAL)
        self.paned_window.pack(fill=tk.BOTH, expand=True)

        # NEW: Figure Setup
        self.fig = Figure(dpi=100)
        self.ax = self.fig.add_subplot(1, 1, 1)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.paned_window)
        self.paned_window.add(self.canvas.get_tk_widget(), weight=4)

        list_container = ttk.Frame(self.paned_window)
        self.paned_window.add(list_container, weight=1)
        
        list_header = ttk.Frame(list_container)
        list_header.pack(fill=tk.X)
        self.log_win_lbl = ttk.Label(list_header, text=t["log_win"])
        self.log_win_lbl.pack(side=tk.LEFT, padx=5)

        #self.data_listbox = tk.Listbox(list_container, selectmode=tk.EXTENDED, font=("Consolas", 9))
        self.data_listbox = tk.Listbox(
            list_container, 
            selectmode=tk.EXTENDED,
            bg="black",           # Μαύρο φόντο
            fg="#00D5FF",         # Neon Green (Πράσινο "Electric")
            selectbackground="#003144",  # Σκούρο πράσινο όταν επιλέγετε μια γραμμή
            selectforeground="white",    # Λευκά γράμματα κατά την επιλογή
            font=("Consolas", 10, "bold") # Monospaced γραμματοσειρά για στυλ τερματικού
        )
        self.data_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scbr = ttk.Scrollbar(list_container, orient=tk.VERTICAL, command=self.data_listbox.yview)
        scbr.pack(side=tk.RIGHT, fill=tk.Y)
        self.data_listbox.config(yscrollcommand=scbr.set)
        
        # --- ΔΕΞΙΑ ΠΛΕΥΡΑ ---
        right_panel = ttk.Frame(main_frame, padding=2)
        right_panel.pack(side=tk.RIGHT, fill=tk.Y, padx=2)

        self.settings_group = ttk.LabelFrame(right_panel, text=" Ρυθμίσεις ", padding=5)
        self.settings_group.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        # Γλώσσα & Οδηγίες
        lang_instr_frame = ttk.Frame(self.settings_group)
        lang_instr_frame.pack(fill=tk.X, pady=1)
        self.lang_btn = ttk.Button(lang_instr_frame, text=t["lang_btn"], command=self.toggle_language)
        self.lang_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=1)
        self.instr_btn = ttk.Button(lang_instr_frame, text=t["instructions"], command=self.open_instructions_window)
        self.instr_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=1)

        # NEW: Επιλογή τύπου διαγράμματος στο UI
        type_frame = ttk.Frame(self.settings_group)
        type_frame.pack(fill=tk.X, pady=5)
        self.type_lbl = ttk.Label(type_frame, text=t["graph_type"])
        self.type_lbl.pack(side=tk.LEFT)
        self.type_combo = ttk.Combobox(type_frame, textvariable=self.graph_type, values=["Linear", "Radar"], state="readonly", width=10)
        self.type_combo.pack(side=tk.LEFT, padx=5)
        self.type_combo.bind("<<ComboboxSelected>>", lambda e: self.clear_data_soft())

        # Port & Baud
        port_frame = ttk.Frame(self.settings_group)
        port_frame.pack(fill=tk.X, pady=8)
        self.port_lbl = ttk.Label(port_frame, text=t["port_label"])
        self.port_lbl.pack(side=tk.LEFT)
        self.ports_combobox = ttk.Combobox(port_frame, state="readonly", width=12)
        self.ports_combobox.pack(side=tk.LEFT, padx=2,pady=(5, 5))
        self.refresh_btn = ttk.Button(port_frame, text="↻", width=3, command=self.refresh_ports)
        self.refresh_btn.pack(side=tk.LEFT)
        self.refresh_ports()

        baud_frame = ttk.Frame(self.settings_group)
        baud_frame.pack(fill=tk.X, pady=1)
        self.baud_lbl = ttk.Label(baud_frame, text=t["baud_label"])
        self.baud_lbl.pack(side=tk.LEFT)
        self.baud_combo = ttk.Combobox(baud_frame, textvariable=self.baudrate, values=[9600, 19200, 38400, 57600, 115200], state="readonly", width=10)
        self.baud_combo.pack(side=tk.LEFT, padx=5)

        ttk.Separator(self.settings_group, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=4)

        # Αρχείο
        self.file_lbl = ttk.Label(self.settings_group, text=t["file_label"])
        self.file_lbl.pack(anchor="w")
        file_row = ttk.Frame(self.settings_group)
        file_row.pack(fill=tk.X)
        self.file_entry = ttk.Entry(file_row, textvariable=self.output_path)
        self.file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.browse_btn = ttk.Button(file_row, text="...", width=3, command=self.browse_file)
        self.browse_btn.pack(side=tk.LEFT, padx=2)

        # Τίτλοι
        self.col_titles_lbl = ttk.Label(self.settings_group, text=t["col_titles"])
        self.col_titles_lbl.pack(anchor="w", pady=(2,0))
        titles_grid = ttk.Frame(self.settings_group)
        titles_grid.pack(fill=tk.X)
        self.extra_text_vars = [tk.StringVar(value=f"Col{i+1}") for i in range(8)]
        for i in range(8):
            r, c = divmod(i, 2)
            ttk.Entry(titles_grid, textvariable=self.extra_text_vars[i], width=9).grid(row=r, column=c, padx=1, pady=1)

        ttk.Separator(self.settings_group, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=4)

        # ThingSpeak
        self.tspeak_chk = ttk.Checkbutton(self.settings_group, text=t["thingspeak"], variable=self.send_to_thingspeak)
        self.tspeak_chk.pack(anchor="w")
        self.ts_api_entry = ttk.Entry(self.settings_group, textvariable=self.thingspeak_api_key)
        self.ts_api_entry.pack(fill=tk.X, pady=1)
        
        ts_int_frame = ttk.Frame(self.settings_group)
        ts_int_frame.pack(fill=tk.X)
        self.ts_interval_lbl = ttk.Label(ts_int_frame, text=t["ts_interval"])
        self.ts_interval_lbl.pack(side=tk.LEFT)
        self.ts_int_entry = ttk.Entry(ts_int_frame, textvariable=self.ts_interval, width=6)
        self.ts_int_entry.pack(side=tk.LEFT, padx=5)

        ttk.Separator(self.settings_group, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=4)

        # Graph Settings
        graph_row1 = ttk.Frame(self.settings_group)
        graph_row1.pack(fill=tk.X)
        self.graph_win_lbl = ttk.Label(graph_row1, text=t["graph_win"])
        self.graph_win_lbl.pack(side=tk.LEFT)
        self.graph_limit_entry = ttk.Entry(graph_row1, textvariable=self.max_val_limit, width=8)
        self.graph_limit_entry.pack(side=tk.LEFT, padx=5)

        graph_row2 = ttk.Frame(self.settings_group)
        graph_row2.pack(fill=tk.X, pady=1)
        self.scroll_chk = ttk.Checkbutton(graph_row2, text=t["scroll"], variable=self.scroll_mode)
        self.scroll_chk.pack(side=tk.LEFT)
        self.scroll_size_entry = ttk.Entry(graph_row2, textvariable=self.scroll_window_size, width=8)
        self.scroll_size_entry.pack(side=tk.LEFT, padx=5)

        # --- ΛΕΙΤΟΥΡΓΙΕΣ (Actions) ---
        self.actions_group = ttk.LabelFrame(right_panel, text=" Actions " if self.current_lang=="EN" else " Λειτουργίες ", padding=5)
        self.actions_group.pack(side=tk.BOTTOM, fill=tk.X, pady=2)
        self.start_btn = ttk.Button(self.actions_group, text=t["start"], command=self.start_logging)
        self.start_btn.grid(row=0, column=0, sticky="ew", padx=1, pady=1, ipady=2)
        self.stop_btn = ttk.Button(self.actions_group, text=t["stop"], command=self.stop_logging)
        self.stop_btn.grid(row=0, column=1, sticky="ew", padx=1, pady=1, ipady=2)
        self.save_btn = ttk.Button(self.actions_group, text=t["save"], command=self.save_data)
        self.save_btn.grid(row=1, column=0, sticky="ew", padx=1, pady=1, ipady=2)
        self.clear_btn = ttk.Button(self.actions_group, text=t["clear"], command=self.clear_data)
        self.clear_btn.grid(row=1, column=1, sticky="ew", padx=1, pady=1, ipady=2)
        self.actions_group.columnconfigure(0, weight=1)
        self.actions_group.columnconfigure(1, weight=1)         

        # Context Menu
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label=t["copy"], command=self.copy_to_clipboard)
        self.context_menu.add_command(label=t["export_csv"], command=self.export_selected_to_csv)
        self.context_menu.add_command(label=t["export_xlsx"], command=self.export_selected_to_xlsx)
        self.data_listbox.bind("<Button-3>", self.show_context_menu)
        self.data_listbox.bind("<Button-2>", self.show_context_menu)
    
    def toggle_language(self):
        self.current_lang = "EN" if self.current_lang == "EL" else "EL"
        t = self.translations[self.current_lang]
        
        self.root.title(t["title"])
        self.title_label.config(text=t["title"])
        self.settings_group.config(text=" Settings " if self.current_lang=="EN" else " Ρυθμίσεις ")
        self.actions_group.config(text=" Actions " if self.current_lang=="EN" else " Λειτουργίες ") 
        
        self.lang_btn.config(text=t["lang_btn"])
        self.instr_btn.config(text=t["instructions"])
        self.start_btn.config(text=t["start"])
        self.stop_btn.config(text=t["stop"])
        self.save_btn.config(text=t["save"])
        self.clear_btn.config(text=t["clear"])
        
        self.port_lbl.config(text=t["port_label"])
        self.baud_lbl.config(text=t["baud_label"])
        self.file_lbl.config(text=t["file_label"])
        self.col_titles_lbl.config(text=t["col_titles"])
        self.tspeak_chk.config(text=t["thingspeak"])
        self.ts_interval_lbl.config(text=t["ts_interval"])
        self.graph_win_lbl.config(text=t["graph_win"])
        self.scroll_chk.config(text=t["scroll"])
        self.log_win_lbl.config(text=t["log_win"])
        self.type_lbl.config(text=t["graph_type"]) # NEW

        self.context_menu.entryconfigure(0, label=t["copy"])
        self.context_menu.entryconfigure(1, label=t["export_csv"])
        self.context_menu.entryconfigure(2, label=t["export_xlsx"])

    def open_instructions_window(self):
        instructions_window = tk.Toplevel(self.root)
        instructions_window.title("Οδηγίες / Instructions")
        instructions_window.geometry("900x700")
        
        text_el = (
            "Καταγραφή δεδομένων από serial (Serial Data Logger).\n\n\n"
            "Μπορείτε να:\n\n" 
            "1. Eπιλέξετε ένα από τα δύο διαγράμματα ΓΡΑΜΙΚΟ (linear) και (Radar)\n"
            "   Α. Γραμικό (linear) απεικόνιση μέχρι και 8 τιμών με διαφορετικο χρώμα\n"
            "   Β. (Radar)απεικόνιση μέχρι και 2 τιμών: Μήκος και Κατεύθυνση\n\n"
            "2. Ορίστε τη θύρα από την οποία θα διαβάσετε δεδομένα.\n"
            "   (με [Aνανέωση] διαβάζονται ξανά οι διαθέσιμες θύρες, \n"
            "   σε περίπτωση που συνδέσατε τον μικροεπεξεργαστή μετά το άνοιγμα αυτής εδώ της εφαρμογής)\n\n"
            "3. Ορίσετε το Baudrate για τη σύνδεση (Παράδειγμα: 9600 για Mind+ ή 115200 για MakeCode).\n\n"
            "4. Επιλέξετε το όνομα του αρχείου και τον τύπο του (.xlsx ή .csv), για αποθήκευση των μετρήσεων.\n\n"
            "5. Ορίστε τους τίτλους των στηλών στο .xlsx (μέχρι 8)\n\n"
            "6. Επιλέξετε αν οι μετρήσεις (μέχρι 8) θα εξάγονται ταυτόχρονα στο ThinkSpeeak το οποίο δέχεται τιμές κάθε 15''.\n"
            "   (Θα χρειαστεί να oρίσετε και το API Key που θα βρείτε στην αντίστοιχη επιλογή της διαδικτυακής εφαρμογής ThinkSpeak).\n\n"
            "7. Επιλέξετε την καθυστέρηση μεταξύ των δειγματοληψιών (καλό είναι να ρυθμίζεται από το πρόγραμμα που τις εξάγει)\n\n"
            "8. Επιλέξετε αν θα κυλίεται το διάγραμμα προς τα αριστερά και για πόσες τιμές\n\n"
            "9. Επιλέξετε κάθε ποσες γραμμές θα διαγράφεται το 10% από το παράθυρο προβολής τιμών\n"
            "   (Στο αρχείο που θα αποθηκεύετε στο τέλος θα είναι όλες οι τιμές  ανεξάρτητα από το πόσες εμφανίζονται στο παράθυρο μετρήσεων\n\n"
            "10. Επιλέξετε το άνω όριο των τιμών που θα εμφανίζονται στο διάγραμμα (όριο άξονα y)\n\n\n\n"
            
            "Λειτουργίες:\n"
            "_______________________\n\n"
            "Πατήστε [Έναρξη] για να ξεκινήσετε τη καταγραφή.\n"
            "Πατήστε [Τερματισμός] για να σταματήσετε την καταγραφή.\n"
            "Πατήστε [Αποθήκευση στο αρχείο] για να αποθηκεύσετε τις μετρήσεις στο αρχείο που ήδη έχετε επιλέξει.\n"
            "   (μπορείτε και πριν τον τερματισμό να αποθηκεύετε τιμές στο αρχείο, οι οποίες θα προστεθούν σ' αυτό)\n"
            "Εναλλακτικά μπορείτε να αποθηκεύσετε στη μνήμη, σε άλλο αρχείο (.xlsx, .csv) ...και με δεξί κλικ πάνω στο παράθυρο των τιμών \n"
            "   (επιλέγοντας κάποιες aαπό αυτές ή/και όλες τις γραμμές που έχουν καταγραφεί).\n\n"
            "Πατήστε [Καθαρισμός] για να καθαρίσετε το διάγραμμα και τις τρέχουσες τιμές\n"
            "   (Δεν διαγράγονται τιμές από το .xlsx που ήδη έχετε αποθηκεύσει από προηγούμενη φορά). \n"
            "Μπορείτε να ρυθμίσετε σε μήκος το παράθυρο καταγραφής τιμών ...και του διαγράμματος,  \n" 
            "   σύρροντας την ενδιάμεση διαχωριστική μπάρα δεξιά ή αριστερά.\n\n\n"
           
          
            "Ελπίζω να σας φανεί χρήσιμη η εφαρμογή αυτή.\n"
        )
        
        text_en = (
            "Serial Data Logger - Data Recording.\n\n\n"
            "You can:\n\n" 
            "1. Select between two chart types: LINEAR and RADAR.\n"
            "   A. Linear: Displays up to 8 values, each with a different color.\n"
            "   B. Radar: Displays up to 2 values: Distance and Direction.\n\n"
            "2. Set the Port to read data from.\n"
            "   (Use [Refresh] to reload available ports if you connected the \n"
            "   microprocessor after opening this application).\n\n"
            "3. Set the Baudrate for the connection (Example: 9600 for Mind+ or 115200 for MakeCode).\n\n"
            "4. Choose the file name and type (.xlsx or .csv) to save your measurements.\n\n"
            "5. Define column titles for the .xlsx file (up to 8).\n\n"
            "6. Choose if measurements (up to 8) will be exported simultaneously to ThingSpeak \n"
            "   (updates every 15''). You will need to provide your API Key.\n\n"
            "7. Set the delay between samples (ideally managed by the source program).\n\n"
            "8. Enable/disable diagram scrolling and set the number of visible points.\n\n"
            "9. Set the line limit for the data log window; once reached, 10% of old entries \n"
            "   are cleared (Note: The final file will contain all values regardless).\n\n"
            "10. Set the upper limit for the values displayed on the chart (Y-axis limit).\n\n\n\n"
            
            "Actions:\n"
            "_______________________\n\n"
            "Press [Start] to begin recording.\n"
            "Press [Stop] to end recording.\n"
            "Press [Save] to store measurements in your selected file.\n"
            "   (Values can be saved during recording and will be appended to the file).\n"
            "Alternatively, you can export specific data by right-clicking in the log window \n"
            "   to copy or export selected lines to .xlsx or .csv.\n\n"
            "Press [Clear] to reset the chart and current session data.\n"
            "   (This does not delete data already saved in your files).\n"
            "You can adjust the height of the log window and chart by dragging \n"
            "   the horizontal separator bar up or down.\n\n\n"
        
            "I hope you find this application useful.\n"
        )
        
        display_text = text_el if self.current_lang == "EL" else text_en
        # Δημιουργία πλαισίου για το κείμενο και την μπάρα κύλισης
        frame = ttk.Frame(instructions_window)
        frame.pack(expand=True, fill="both", padx=10, pady=10)
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        # font=("Arial", 14) 
        text_widget = tk.Text(frame, wrap=tk.WORD, font=("Arial", 14), 
                      yscrollcommand=scrollbar.set, 
                      bg="#E8E3E3", 
                      fg="#00008B",  
                      relief="flat")
        text_widget.insert(tk.END, display_text)
        text_widget.config(state=tk.DISABLED) # Για να μην μπορεί να το σβήσει ο χρήστης
        text_widget.pack(side=tk.LEFT, expand=True, fill="both")
        scrollbar.config(command=text_widget.yview)
        ttk.Button(instructions_window, text="OK", command=instructions_window.destroy).pack(pady=10)

    def get_text(self, key):
        return self.translations[self.current_lang].get(key, key)

    def refresh_ports(self):
        ports = [port.device for port in list_ports.comports()]
        self.ports_combobox["values"] = ports
        if ports: self.ports_combobox.current(0)

    def browse_file(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")])
        if file_path: self.output_path.set(file_path)

    def show_context_menu(self, event):
        self.context_menu.tk_popup(event.x_root, event.y_root)

    def copy_to_clipboard(self, event=None):
        selected_indices = self.data_listbox.curselection()
        if not selected_indices: return
        selected_text = "\n".join([self.data_listbox.get(i) for i in selected_indices])
        self.root.clipboard_clear()
        self.root.clipboard_append(selected_text)

    def start_logging(self):
        try:
            self.serial_port = serial.Serial(self.ports_combobox.get(), baudrate=self.baudrate.get(), timeout=1)
            self.stop_event.clear()
            threading.Thread(target=self.record_data, daemon=True).start()
            self.update_plot()
        except Exception as e: messagebox.showerror("Connection Error", str(e))

    def stop_logging(self):
        if self.serial_port:
            self.stop_event.set()
            self.serial_port.close()
            self.serial_port = None

    def record_data(self):
        try:
            while not self.stop_event.is_set():
                line = self.serial_port.readline().decode('utf-8', errors='ignore').strip() 
                if line:
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    current_max = self.max_val_limit.get()
                    line = line.replace(';', ',').replace(':', ',')
                    raw_items = [item.strip() for item in line.split(',') if item.strip()]
                    clean_numeric_values = []
                    for item in raw_items:
                        try:
                            val = float(item)
                            clean_numeric_values.append(0.0 if val > current_max else val)
                        except ValueError: clean_numeric_values.append(0.0)
                    self.data_queue.put((timestamp, clean_numeric_values, raw_items))
                    self.send_to_thingspeak_api(clean_numeric_values)
                    threading.Event().wait(self.sampling_rate.get() / 1000)
        except Exception as e:
            if not self.stop_event.is_set(): messagebox.showerror("Error", str(e))

 
    def update_plot(self):
        # 1. Λήψη όλων των διαθέσιμων δεδομένων από την ουρά
        while not self.data_queue.empty():
            timestamp, numeric_values, raw_items = self.data_queue.get()
            self.times.append(len(self.times) + 1)
            self.actual_timestamps.append(timestamp)
            self.values.append(numeric_values)
            
            # Ενημέρωση Listbox
            self.data_listbox.insert(tk.END, f"{timestamp}: {', '.join(raw_items)}")
            self.data_listbox.see(tk.END)
            if self.data_listbox.size() > self.listbox_limit.get():
                self.data_listbox.delete(0, int(self.listbox_limit.get() * 0.1))

        if self.times:
            is_radar = (self.graph_type.get() == "Radar")
            
            # 2. ΑΥΤΟΜΑΤΗ ΔΙΑΧΕΙΡΙΣΗ ΠΑΡΑΘΥΡΟΥ ΚΥΛΙΣΗΣ
            # Στο Radar κρατάμε μόνο την τελευταία τιμή για ταχύτητα, στο Linear όσες ορίζει το UI
            if is_radar:
                current_window = 1
            else:
                current_window = self.scroll_window_size.get() if self.scroll_mode.get() else len(self.times)

            # 3. Έλεγχος αν πρέπει να αλλάξουμε τύπο γραφήματος (Polar vs Cartesian)
            current_is_polar = hasattr(self.ax, 'set_theta_zero_location')
            if is_radar != current_is_polar:
                self.fig.clear()
                if is_radar:
                    self.ax = self.fig.add_subplot(1, 1, 1, polar=True)
                else:
                    self.ax = self.fig.add_subplot(1, 1, 1)
            
            self.ax.clear()

            if is_radar:
                # --- RADAR / COMPASS LOGIC (DARK MODE) ---
                plot_values = self.values[-current_window:]
                if plot_values and len(plot_values[-1]) >= 2:
                    dist = plot_values[-1][0]
                    angle_deg = plot_values[-1][1]
                    angle_rad = np.deg2rad(angle_deg)
                    
                    # Ρυθμίσεις Χρωμάτων & Φόντου
                    self.fig.set_facecolor('black')
                    self.ax.set_facecolor("#033403") # Πολύ σκούρο πράσινο
                    
                    # Ρυθμίσεις Προσανατολισμού & Ορίων
                    self.ax.set_theta_zero_location('N') # 0° πάνω
                    self.ax.set_theta_direction(-1)      # Δεξιόστροφα
                    self.ax.set_thetalim(0, 2*np.pi)      # Κλείδωμα 360 μοιρών
                    self.ax.set_rmax(self.max_val_limit.get())
                    
                    # Πλέγμα και Ενδείξεις
                    self.ax.grid(True, color="#31AC31", linestyle='--') # Σκούρο πράσινο πλέγμα
                    self.ax.tick_params(colors='white')                 # Λευκοί αριθμοί απόστασης
                    
                    # Ετικέτες Πυξίδας (N, E, S, W)
                    self.ax.set_thetagrids([0, 45, 90, 135, 180, 225, 270, 315], 
                                          ['N', 'NE', 'E', 'SE', 'S', 'SW', 'W', 'NW'],
                                          fontsize=10, fontweight='bold', color='white')
                    
                    # Σχεδίαση Πράσινης Βελόνας
                    self.ax.plot([angle_rad, angle_rad], [0, dist], color='#00FF00', lw=5)
                    self.ax.scatter(angle_rad, dist, color='#00FF00', s=100, 
                                    edgecolors='white', linewidth=1, zorder=5)
                    
                    # Τίτλος (Heading)
                    self.ax.set_title(f"{self.get_text('heading')}: {angle_deg}°", color='#00FF00', 
                                    fontsize=12, fontweight='bold', pad=20)
                    
            else:
                # --- LINEAR LOGIC (NORMAL MODE) ---
                self.fig.set_facecolor('#F0F0F0') # Επαναφορά στο αρχικό χρώμα
                self.ax.set_facecolor('white')
                self.ax.tick_params(colors='black', labelcolor='black')
                self.ax.xaxis.label.set_color('black')
                self.ax.yaxis.label.set_color('black')
                
                plot_times = self.times[-current_window:]
                plot_values = self.values[-current_window:]
                
                # Ομαδοποίηση δεδομένων ανά στήλη
                data_cols = list(zip_longest(*plot_values, fillvalue=0.0))
                for i, col in enumerate(data_cols[:8]):
                    label_name = self.extra_text_vars[i].get() or f"Val {i+1}"
                    self.ax.plot(plot_times, list(col), label=label_name)
                
               
                # Δυναμική μετάφραση αξόνων
                x_text = self.get_text('x_label')
                if self.scroll_mode.get():
                    points_val = self.scroll_window_size.get()
                    x_text += f" ({points_val} {self.get_text('last_points')})"
                    
                self.ax.set_xlabel(x_text, color='black')
                self.ax.set_ylabel(self.get_text('y_label'), color='black')
          
                self.ax.legend(loc='upper left', fontsize='small')
                self.ax.set_title(self.get_text('linear_title'), color='black')

            self.canvas.draw()

        # Επαναπρογραμματισμός της επόμενης ανανέωσης
        if not self.stop_event.is_set():
            self.root.after(200, self.update_plot)












    def clear_data_soft(self):
        # NEW: Βοηθητική για καθαρισμό μόνο του γραφήματος κατά την εναλλαγή mode
        self.ax.clear()
        self.canvas.draw()

  
    def send_to_thingspeak_api(self, values):
        if self.send_to_thingspeak.get():
            now = datetime.now()
            # Υπολογισμός διαφοράς χρόνου σε δευτερόλεπτα
            diff = (now - self.last_ts_send).total_seconds()
            
            if diff >= self.ts_interval.get():
                self.last_ts_send = now
                threading.Thread(target=self._async_ts, args=(values,), daemon=True).start()

    def _async_ts(self, values):
        try:
            url = "https://api.thingspeak.com/update"
            params = {"api_key": self.thingspeak_api_key.get()}
            for i, v in enumerate(values[:8]): params[f"field{i+1}"] = v
            requests.get(url, params=params, timeout=5)
        except: pass

    def save_data(self):
        path = self.output_path.get()
        if not self.times: return
        headers = ["Time"] + [v.get() for v in self.extra_text_vars if v.get()]
        rows = [[t] + list(v) for t, v in zip(self.actual_timestamps, self.values)]
        try:
            if path.endswith(".xlsx"):
                wb = openpyxl.load_workbook(path) if os.path.exists(path) else Workbook()
                ws = wb.active
                if wb.sheetnames == ['Sheet']: ws.append(headers)
                for r in rows: ws.append(r)
                wb.save(path)
            else:
                with open(path, "a", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    if not os.path.exists(path): writer.writerow(headers)
                    writer.writerows(rows)
            messagebox.showinfo("Save", "Success!")
        except Exception as e: messagebox.showerror("Error", str(e))

    def clear_data(self):
        if messagebox.askyesno("Clear", "Delete all data?"):
            self.times, self.values, self.actual_timestamps = [], [], []
            self.data_listbox.delete(0, tk.END)
            self.ax.clear()
            self.canvas.draw()

    def export_selected_to_csv(self): self._export_selected_logic(".csv")
    def export_selected_to_xlsx(self): self._export_selected_logic(".xlsx")
    
    def _export_selected_logic(self, extension):
        selected_indices = self.data_listbox.curselection()
        if not selected_indices: return
        file_path = filedialog.asksaveasfilename(defaultextension=extension)
        if not file_path: return
        try:
            headers = ["Timestamp"] + [v.get() for v in self.extra_text_vars if v.get()]
            rows = []
            for i in selected_indices:
                raw_line = self.data_listbox.get(i)
                if ": " in raw_line:
                    ts, vals = raw_line.split(": ", 1)
                    rows.append([ts] + vals.split(", "))
            if extension == ".csv":
                with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                    csv.writer(f).writerow(headers)
                    csv.writer(f).writerows(rows)
            else:
                nb = Workbook()
                ws = nb.active
                ws.append(headers)
                for r in rows: ws.append(r)
                nb.save(file_path)
            messagebox.showinfo("Export", "Done!")
        except Exception as e: messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = SerialDataLogger(root)
    root.mainloop()
